# 接口导出（样式化 Excel）相关方法
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from .models import InterfaceInfo, InterfaceField

DefaultStyle={
    "font" : Font(name='Calibri',
                    size=11,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000'),
    'fill' : PatternFill(patternType='solid',fgColor="99CCFF"),
    'border' : Border(left=Side(border_style='thin', color='FF000000'),
                    right=Side(border_style='thin',color='FF000000'),
                    top=Side(border_style='thin',color='FF000000'),
                    bottom=Side(border_style='thin',color='FF000000')
                )
}

def set_area_border(ws,start_row,end_row,start_col,end_col):
    for row in range(start_row,end_row+1):
        for col in range(start_col,end_col+1):
            ws.cell(row=row,column=col).border = DefaultStyle['border']

def build_interface_workbook_bytes(interface: InterfaceInfo, fields) -> bytes:
    """构建 Excel 并返回二进制内容"""
    wb = make_interface_workbook(interface, fields)
    import io
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def make_interface_workbook(interface: InterfaceInfo, fields):
    """
    生成样式化的接口导出工作簿：
    - 顶部信息区（名称、编码、各开关、数据库信息、SQL 等）
    - 字段列表（按输入参数、输出参数顺序）
    返回 openpyxl.Workbook
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "report"

    template_info = [
        {'position': 'A1', 'name': '报表平台'},
        {'position': 'C1', 'name': '模块名称'},
        {'position': 'E1', 'name': '报表名称'},
        {'position': 'G1', 'name': '报表代码'},
        {'position': 'A2', 'name': '接口名称'},
        {'position': 'C2', 'name': '接口代码'},
        {'position': 'E2', 'name': '日期选项'},
        {'position': 'G2', 'name': '二级表头'},
        {'position': 'I2', 'name': '需要登录'},
        {'position': 'K2', 'name': '告警方式'},
        {'position': 'M2', 'name': '接口状态'},
        {'position': 'A3', 'name': '数据库类型'},
        {'position': 'C3', 'name': '数据库名称'},
        {'position': 'E3', 'name': '接口sql'},
        {'position': 'A4', 'name': '是否分页'},
        {'position': 'C4', 'name': '是否合计'},
        {'position': 'E4', 'name': '合计sql'},
    ]
    for info in template_info:
        cell = ws[info['position']]
        cell.value = info['name']
        cell.font = DefaultStyle['font']
        cell.fill = DefaultStyle['fill']
        cell.border = DefaultStyle['border']

    # 顶部信息填充（报告归属信息）
    ws['B1'] = interface.platform_name or ''
    ws['D1'] = interface.module_name or ''
    ws['F1'] = interface.report_name or ''
    ws['H1'] = interface.report_code or ''

    ws['B2'] = interface.interface_name
    ws['D2'] = interface.interface_code
    ws['F2'] = '是' if interface.is_date_option == '1' else '否'
    ws['H2'] = '是' if interface.is_second_table == '1' else '否'
    ws['J2'] = '是' if interface.is_login_visit == '1' else '否'
    ws['N2'] = interface.get_enable_display()
    # 报警类型显示值
    try:
        ws['L2'] = interface.get_alarm_type_display()
    except Exception:
        ws['L2'] = interface.alarm_type

    ws['B3'] = interface.interface_db_type
    ws['D3'] = interface.interface_db_name
    ws['F3'] = interface.interface_sql or ''
    ws['B4'] = '是' if interface.is_paging == '1' else '否'
    ws['D4'] = '是' if interface.is_total == '1' else '否'
    ws['F4'] = interface.total_sql or ''

    # 列头
    column_headers = [
        {'label': '序号'},
        {'label': '参数名称'},
        {'label': '参数代码'},
        {'label': '参数类型'},
        {'label': '数据类型'},
        {'label': '是否展示'},
        {'label': '是否导出'},
        {'label': '参数接口代码'},
        {'label': '参数默认值'},
        {'label': '级联参数'},
        {'label': '父表头名称'},
        {'label': '父表头位置'},
        {'label': '是否合并行'},
        {'label': '是否显示备注'},
        {'label': '参数描述'},
    ]
    for index, column_info in enumerate(column_headers):
        cell = ws.cell(row=5, column=index + 1)
        cell.font = DefaultStyle['font']
        cell.fill = PatternFill(patternType='solid', fgColor='C0C0C0')
        cell.value = column_info.get('label')

    # 字段列表：输入参数在前，输出参数在后
    fields_sorted = list(fields) if fields is not None else []
    fields_sorted.sort(key=lambda x: (x.interface_para_type, x.interface_para_position))
    fields_type_input = [f for f in fields_sorted if getattr(f, 'interface_para_type', '') == '1']
    fields_type_output = [f for f in fields_sorted if getattr(f, 'interface_para_type', '') == '2']
    ordered_fields = fields_type_input + fields_type_output

    for row_index, f in enumerate(ordered_fields):
        # 逐列写入
        values = [
            f.interface_para_position,
            f.interface_para_name,
            f.interface_para_code,
            getattr(f, 'get_interface_para_type_display', lambda: f.interface_para_type)(),
            getattr(f, 'get_interface_data_type_display', lambda: f.interface_data_type)(),
            # '是' if f.interface_show_flag == '1' else '否',
            getattr(f, 'get_interface_show_flag_display', lambda: f.interface_show_flag)(),
            # '是' if f.interface_export_flag == '1' else '否',
            getattr(f, 'get_interface_show_flag_display', lambda: f.interface_show_flag)(),
            f.interface_para_interface_code or '',
            f.interface_para_default or '',
            f.interface_cascade_para or '',
            f.interface_parent_name or '',
            f.interface_parent_position or '0',
            # 跨行显示
            getattr(f, 'get_interface_para_rowspan_display', lambda: f.interface_para_rowspan)(),
            getattr(f, 'get_interface_show_desc_display', lambda: f.interface_show_desc)(),
            getattr(f, 'get_interface_para_desc_display', lambda: f.interface_para_desc)(),
        ]
        for col_index, val in enumerate(values):
            ws.cell(row=row_index + 6, column=1 + col_index).value = val

    set_area_border(ws, ws.min_row, ws.max_row, ws.min_column, ws.max_column)
    ws['P1'] = 'report'
    return wb

def parse_interface_workbook(wb: Workbook) -> list[tuple[InterfaceInfo, list[InterfaceField]]]:
    """
    解析 Excel 中的接口定义（基本信息 + 字段列表）
    支持批量导入：每个 Sheet 解析为一个接口
    """
    results = []
    
    # 辅助函数：解析 Yes/No
    def parse_yes_no(val):
        if not val:
            return '0'
        return '1' if '是' in str(val) else '0'
        
    def parse_enable(val):
        if not val:
            return '1'
        return '1' if '启用' in str(val) else '0'
    # 辅助函数：解析报警类型
    def parse_alarm_type(val):
        if not val:
            return '0'
        val_str = str(val)
        # 映射表：根据 InterfaceInfo.ALARM_TYPE_CHOICES
        mapping = {
            '否': '0', '邮件': '1', '短信': '2', '钉钉': '3', 
            '企业微信': '4', '电话': '5', '飞书': '6'
        }
        return mapping.get(val_str, '0')

    # 辅助函数：解析参数类型
    def parse_para_type(val):
        if not val:
            return '1' # 默认输入参数
        return '2' if '输出' in str(val) else '1'

    # 辅助函数：解析数据类型
    def parse_data_type(val):
        if not val:
            return '15' # 默认文本
        # 映射表 InterfaceField.DATA_TYPE_CHOICES
        mapping = {
            '字符': '1', '整数': '2', '小数': '3', '百分比': '4',
            '无格式整数': '5', '无格式小数': '6', '无格式百分比': '7',
            '1位百分比': '8', '1位小数': '9', '年份': '10',
            '日期': '11', '月份': '12', '单选': '13', '多选': '14',
            '文本': '15'
        }
        return mapping.get(str(val), '15')

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # 简单校验：检查 A1 是否为 "报表平台"
        if ws['A1'].value != '报表平台':
            continue

        # 1. 解析基本信息
        info = InterfaceInfo()
        # B1: 平台名称
        info.platform_name = ws['B1'].value
        # D1: 模块信息
        info.module_name = ws['D1'].value
        # F1
        info.report_name = ws['F1'].value
        # H1
        info.report_code = ws['H1'].value
        # B2: 接口名称
        info.interface_name = ws['B2'].value
        # D2: 接口代码
        info.interface_code = ws['D2'].value
        
        # 如果没有名称或代码，跳过
        if not info.interface_name or not info.interface_code:
            continue
            
        # F2: 日期选项
        info.is_date_option = parse_yes_no(ws['F2'].value)
        # H2: 二级表头
        info.is_second_table = parse_yes_no(ws['H2'].value)
        # J2: 需要登录
        info.is_login_visit = parse_yes_no(ws['J2'].value)
        # L2: 告警方式
        info.alarm_type = parse_alarm_type(ws['L2'].value)
        # N2: 接口状态
        info.enable = parse_enable(ws['N2'].value)
        # B3: 数据库类型
        info.interface_db_type = ws['B3'].value or 'mysql'
        # D3: 数据库名称
        info.interface_db_name = ws['D3'].value or ''
        # F3: 接口sql
        info.interface_sql = ws['F3'].value or ''
        
        # B4: 是否分页
        info.is_paging = parse_yes_no(ws['B4'].value)
        # D4: 是否合计
        info.is_total = parse_yes_no(ws['D4'].value)
        # F4: 合计sql
        info.total_sql = ws['F4'].value or ''
        
        # 2. 解析字段列表 (从第 6 行开始)
        fields = []
        row_idx = 6
        max_row = ws.max_row
        
        while row_idx <= max_row:
            # 如果没有参数名称(B列)和参数代码(C列)，视为结束或空行
            para_name = ws.cell(row=row_idx, column=2).value
            para_code = ws.cell(row=row_idx, column=3).value
            
            if not para_name and not para_code:
                row_idx += 1
                continue
                
            field = InterfaceField()
            # A列: 序号/位置 -> interface_para_position
            pos_val = ws.cell(row=row_idx, column=1).value
            try:
                field.interface_para_position = int(pos_val)
            except:
                field.interface_para_position = (row_idx - 5) * 10
                
            field.interface_para_name = para_name or ''
            field.interface_para_code = para_code or ''
            
            # D列: 参数类型
            field.interface_para_type = parse_para_type(ws.cell(row=row_idx, column=4).value)
            # E列: 数据类型
            field.interface_data_type = parse_data_type(ws.cell(row=row_idx, column=5).value)
            # F列: 是否展示
            field.interface_show_flag = parse_yes_no(ws.cell(row=row_idx, column=6).value)
            # G列: 是否导出
            field.interface_export_flag = parse_yes_no(ws.cell(row=row_idx, column=7).value)
            
            # H列: 参数接口代码
            field.interface_para_interface_code = ws.cell(row=row_idx, column=8).value
            # I列: 参数默认值
            val = ws.cell(row=row_idx, column=9).value
            field.interface_para_default = str(val) if val is not None else ''
            # J列: 级联参数
            field.interface_cascade_para = ws.cell(row=row_idx, column=10).value
            # K列: 父表头名称
            field.interface_parent_name = ws.cell(row=row_idx, column=11).value
            # L列: 父表头位置
            p_pos = ws.cell(row=row_idx, column=12).value
            try:
                field.interface_parent_position = int(p_pos) if p_pos is not None else 0
            except:
                field.interface_parent_position = 0
                
            # M列: 是否合并行 (rowspan) -> 这里逻辑有点特殊，model里是 1=是, 0=否
            # 但 excel 可能是 '是'/'否'
            rowspan_val = ws.cell(row=row_idx, column=13).value
            if rowspan_val == '是' or rowspan_val == 1:
                field.interface_para_rowspan = 1
            else:
                field.interface_para_rowspan = 0
                
            # N列: 是否显示备注 -> interface_show_desc (model definition seems to imply show_flag type choice, but field name suggests desc)
            # Checking model: interface_show_desc = models.CharField(..., choices=SHOW_FLAG_CHOICES)
            field.interface_show_desc = parse_yes_no(ws.cell(row=row_idx, column=14).value)
            
            # O列: 参数描述
            field.interface_para_desc = ws.cell(row=row_idx, column=15).value
            
            fields.append(field)
            row_idx += 1
        
        results.append((info, fields))
        
    return results


